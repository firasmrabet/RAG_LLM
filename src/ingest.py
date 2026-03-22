"""
ingest.py — Phase 1 : Lecture et extraction des documents FallahTech
Lit tous les PDFs et le fichier Excel du dossier /documents
et sauvegarde le texte brut dans chroma_db/raw_docs.json
"""

import json
from pathlib import Path
from pypdf import PdfReader
from openpyxl import load_workbook

# ── Chemins ──────────────────────────────────────────────
ROOT      = Path(__file__).parent.parent
DOCS_DIR  = ROOT / "documents"
OUT_PATH  = ROOT / "chroma_db" / "raw_docs.json"

# Fichiers à ignorer
EXCLUDE = {
    "Enonce_MiniProjet_FallahTech.docx.pdf",
    "0_0_Index_DataRoom.pdf"
}

# ── Extraction PDF ────────────────────────────────────────
def extract_pdf(path: Path) -> str:
    reader = PdfReader(path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"[PAGE {i+1}]\n{text}")
    return "\n\n".join(pages)

# ── Extraction Excel ──────────────────────────────────────
def extract_xlsx(path: Path) -> str:
    wb = load_workbook(path, data_only=True)
    all_text = []
    for sheet in wb.worksheets:
        all_text.append(f"[SHEET: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(
                str(c) for c in row if c is not None
            )
            if row_text.strip():
                all_text.append(row_text)
    return "\n".join(all_text)

# ── Ingestion principale ──────────────────────────────────
def ingest():
    if not DOCS_DIR.exists():
        print(f"❌ Dossier introuvable : {DOCS_DIR}")
        return

    docs = []
    print("📂 Lecture des documents...\n")

    for file in sorted(DOCS_DIR.iterdir()):
        if file.name in EXCLUDE or not file.is_file():
            print(f"  ⊘  {file.name} (ignoré)")
            continue

        try:
            if file.suffix.lower() == ".pdf":
                text = extract_pdf(file)
            elif file.suffix.lower() in (".xlsx", ".xlsm"):
                text = extract_xlsx(file)
            else:
                continue

            if text.strip():
                docs.append({
                    "id":     file.name,
                    "source": file.name,
                    "text":   text
                })
                print(f"  ✅ {file.name} "
                      f"({len(text):,} caractères)")
            else:
                print(f"  ⚠️  {file.name} (vide)")

        except Exception as e:
            print(f"  ❌ {file.name} → {e}")

    # Sauvegarde JSON
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(docs, f, ensure_ascii=False, indent=2)

    print(f"\n✅ {len(docs)} documents sauvegardés → {OUT_PATH}")

if __name__ == "__main__":
    ingest()
